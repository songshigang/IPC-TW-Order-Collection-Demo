import openpyxl
wb = openpyxl.load_workbook(r'D:\德亦信息\金佰利\IPC\TW\订单自动化收集平台\RFP\主数据例子.xlsx')

# Get material data
ws = wb['物料']
materials = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[0] and row[1]:
        code = str(row[0]).strip()
        name = str(row[1]).strip()
        box_qty = int(row[2]) if row[2] else 1
        # Determine category
        if '好奇' in name:
            category = '嬰幼兒護理'
            brand = 'Huggies'
        elif '高洁丝' in name:
            category = '女性護理'
            brand = 'Kotex'
        elif '得伴' in name:
            category = '成人護理'
            brand = 'Depend'
        elif '湿巾' in name or '棉柔巾' in name or '植物柔巾' in name:
            category = '紙巾'
            brand = 'Kleenex'
        else:
            category = '紙巾'
            brand = 'Kleenex'
        materials.append({'code': code, 'name': name, 'boxQty': box_qty, 'category': category, 'brand': brand})

# Take first 200
selected = materials[:200]

# Generate JavaScript format
output = []
for m in selected:
    # Extract spec from name - take part after first space or first 20 chars
    parts = m['name'].split(' ')
    if len(parts) > 1:
        spec = parts[1] if len(parts[1]) <= 30 else parts[1][:30]
    else:
        spec = m['name'][:30]

    price = 850 if m['category'] == '嬰幼兒護理' else 320 if m['category'] == '女性護理' else 450 if m['category'] == '成人護理' else 399

    line = "{ productCode: '" + m['code'] + "', productName: '" + m['name'] + "', productLine: '" + m['category'] + "', brand: '" + m['brand'] + "', specification: '" + spec + "', baseUnit: 'EA', boxQty: " + str(m['boxQty']) + ", conversionRate: 1, unitPrice: 'TW$ " + str(price) + "', status: 1 }"
    output.append(line)

# Write to file
with open(r'd:\德亦信息\金佰利\IPC\TW\订单自动化收集平台\IPC TW Order Collection Demo 1.0\products_data.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(output))

print('Generated', len(output), 'product records')
